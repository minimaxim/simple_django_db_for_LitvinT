import csv
from datetime import datetime

import openpyxl
from django.contrib import admin
from django.http import HttpResponse

from .models import User, Company


@admin.action(description='Выгрузить Excel')
def create_excel_user(self, request, queryset):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    headers = ['name', 'country', 'email', 'phone', 'login_bitmain', 'telegram_link', 'instagram_link', 'twitter_link',
               'vk_link', 'facebook_link', 'linkedin_link', 'whatsapp_link', 'counter', 'feedback']
    for col_num, column_title in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    for row_num, obj in enumerate(queryset, 2):
        worksheet.cell(row=row_num, column=1).value = obj.name
        worksheet.cell(row=row_num, column=2).value = obj.country
        worksheet.cell(row=row_num, column=3).value = obj.email
        worksheet.cell(row=row_num, column=4).value = obj.phone
        worksheet.cell(row=row_num, column=5).value = obj.login_bitmain
        worksheet.cell(row=row_num, column=6).value = obj.telegram_link
        worksheet.cell(row=row_num, column=7).value = obj.instagram_link
        worksheet.cell(row=row_num, column=8).value = obj.twitter_link
        worksheet.cell(row=row_num, column=9).value = obj.vk_link
        worksheet.cell(row=row_num, column=10).value = obj.facebook_link
        worksheet.cell(row=row_num, column=11).value = obj.linkedin_link
        worksheet.cell(row=row_num, column=12).value = obj.whatsapp_link
        worksheet.cell(row=row_num, column=13).value = obj.counter
        worksheet.cell(row=row_num, column=14).value = obj.feedback

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=export{datetime.strftime(datetime.now(), "%d/%m/%y")}.xlsx'
    workbook.save(response)

    return response


@admin.action(description='Выгрузить Excel')
def create_excel_company(self, request, queryset):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    headers = ['name', 'country', 'email', 'phone', 'individual', 'individual2', 'telegram_link', 'instagram_link',
               'twitter_link', 'vk_link', 'facebook_link', 'linkedin_link', 'whatsapp_link', 'counter', 'feedback']
    for col_num, column_title in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    for row_num, obj in enumerate(queryset, 2):
        worksheet.cell(row=row_num, column=1).value = obj.name
        worksheet.cell(row=row_num, column=2).value = obj.country
        worksheet.cell(row=row_num, column=3).value = obj.email
        worksheet.cell(row=row_num, column=4).value = obj.phone
        worksheet.cell(row=row_num, column=5).value = obj.individual
        worksheet.cell(row=row_num, column=5).value = obj.individual2
        worksheet.cell(row=row_num, column=6).value = obj.telegram_link
        worksheet.cell(row=row_num, column=7).value = obj.instagram_link
        worksheet.cell(row=row_num, column=8).value = obj.twitter_link
        worksheet.cell(row=row_num, column=9).value = obj.vk_link
        worksheet.cell(row=row_num, column=10).value = obj.facebook_link
        worksheet.cell(row=row_num, column=11).value = obj.linkedin_link
        worksheet.cell(row=row_num, column=12).value = obj.whatsapp_link
        worksheet.cell(row=row_num, column=13).value = obj.counter
        worksheet.cell(row=row_num, column=14).value = obj.feedback

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response[
        'Content-Disposition'] = f'attachment; filename=export {datetime.strftime(datetime.now(), "%d/%m/%y")}.xlsx'
    workbook.save(response)

    return response


@admin.action(description='Выгрузить CSV')
def create_csv_user(self, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response[
        'Content-Disposition'] = f'attachment; filename="export {datetime.strftime(datetime.now(), "%d/%m/%y")}.csv"'
    writer = csv.writer(response)

    fields = ['name', 'country', 'email', 'phone', 'login_bitmain', 'telegram_link', 'instagram_link',
              'twitter_link', 'vk_link', 'facebook_link', 'linkedin_link', 'whatsapp_link', 'counter', 'feedback']
    writer.writerow(fields)

    for obj in queryset:
        data = [obj.name, obj.country, obj.email, obj.phone, obj.login_bitmain, obj.telegram_link, obj.instagram_link,
                obj.twitter_link, obj.vk_link, obj.facebook_link, obj.linkedin_link, obj.whatsapp_link, obj.counter,
                obj.feedback]
        writer.writerow(data)

    return response


@admin.action(description='Выгрузить CSV')
def create_csv_company(self, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response[
        'Content-Disposition'] = f'attachment; filename="export{datetime.strftime(datetime.now(), "%d/%m/%y")}.csv"'
    writer = csv.writer(response)

    # Здесь указываем необходимые поля модели
    fields = ['name', 'country', 'email', 'phone', 'individual', 'individual2' 'telegram_link', 'instagram_link',
              'twitter_link', 'vk_link', 'facebook_link', 'linkedin_link', 'whatsapp_link', 'counter', 'feedback']
    writer.writerow(fields)

    for obj in queryset:
        # Здесь указываем значения полей модели в нужном порядке
        data = [obj.name, obj.country, obj.email, obj.phone, obj.individual, obj.individual2, obj.telegram_link,
                obj.instagram_link, obj.twitter_link, obj.vk_link, obj.facebook_link, obj.linkedin_link,
                obj.whatsapp_link, obj.counter, obj.feedback]
        writer.writerow(data)

    return response


@admin.register(User)
class UserAdmin(admin.ModelAdmin):
    list_display = (
        'id', 'name', 'country', 'email', 'phone', 'login_bitmain', 'telegram_link', 'instagram_link', 'twitter_link',
        'vk_link', 'facebook_link', 'linkedin_link', 'whatsapp_link', 'counter', 'feedback'
    )
    list_filter = (
        'telegram_link', 'instagram_link', 'twitter_link', 'vk_link', 'facebook_link', 'linkedin_link', 'whatsapp_link',
        'country', 'counter'
    )
    search_fields = ('id',)
    actions = (create_excel_user, create_csv_user)
    list_editable = ('counter', 'feedback')


@admin.register(Company)
class CompanyAdmin(admin.ModelAdmin):
    list_display = (
        'id', 'name', 'country', 'email', 'phone', 'individual', 'individual2', 'telegram_link', 'instagram_link',
        'twitter_link', 'vk_link', 'facebook_link', 'linkedin_link', 'whatsapp_link', 'counter', 'feedback'
    )
    list_filter = (
        'telegram_link', 'instagram_link', 'country', 'twitter_link', 'vk_link', 'facebook_link', 'linkedin_link',
        'whatsapp_link', 'counter'
    )
    search_fields = ('id',)
    actions = (create_excel_company, create_csv_company)
    list_editable = ('counter', 'feedback')
